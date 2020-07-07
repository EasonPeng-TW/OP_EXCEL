import op_advance2
import pandas as pd
import requests
import datetime as dt
from datetime import timedelta
from bs4 import BeautifulSoup
import time
import datetime
import matplotlib.pyplot as plt
from matplotlib.font_manager import findfont, FontProperties
import openpyxl
op = op_advance2.Op()


def analysis_op_excel():
    get_days = int(input('請輸入OP回測天數: '))
    i = 0
    op_data = []
    while True:
        try:
            day_delta = i
            i += 1
            today_to_past = (dt.datetime.now() - timedelta(days= int(day_delta)))
            time.sleep(1)
            true_format = today_to_past.strftime("%Y/%m/%d")

            table, search_web, search_date = op.crawl(search_web='cnp_three', search_date=true_format)
            print('Connecting to web...')
            sfbc_nb = int(table[2].iloc[5][10]) #自營商 BC BP SC SP
            sfbp_nb = int(table[2].iloc[8][10])
            sfsc_nb = int(table[2].iloc[5][12])
            sfsp_nb = int(table[2].iloc[8][12])

            bc_nb = int(table[2].iloc[7][10]) #外資 BC BP SC SP
            bp_nb = int(table[2].iloc[10][10])
            sc_nb = int(table[2].iloc[7][12])
            sp_nb = int(table[2].iloc[10][12])
            # return table, bc_nb, bp_nb, sc_nb, sp_nb, sfbc_nb, sfbp_nb, sfsc_nb, sfsp_nb, search_date
                           
            print(true_format + ' 寫入檔案...')
            op_data.append([search_date, sfbc_nb, sfbp_nb, sfsc_nb, sfsp_nb, bc_nb, bp_nb, sc_nb, sp_nb])
            if i == get_days:
                workbook = openpyxl.load_workbook('test.xlsx')
                sheet = workbook.worksheets[0]
                # if sheet['A1'].value == '自營bc', '自營bp', '自營sc', '自營sp', '外資bc', '外資bp', '外資sc', '外資sp'])
                for data in op_data:
                    sheet.append(data)                    
                workbook.save('test.xlsx')
                break
        except:
            print(true_format)
            print('可能是假日或沒開盤')
            if i == get_days:
                workbook = openpyxl.load_workbook('test.xlsx')
                sheet = workbook.worksheets[0]
                # sheet.append(['日期', '自營bc', '自營bp', '自營sc', '自營sp', '外資bc', '外資bp', '外資sc', '外資sp'])
                for data in op_data:
                    sheet.append(data)
                workbook.save('test.xlsx')
                break

def analysis_fu_excel():
    get_days = int(input('請輸入FU回測天數: '))
    i = 0
    op_data = []
    while True:
        try:
            day_delta = i
            i += 1
            today_to_past = (dt.datetime.now() - timedelta(days= int(day_delta)))
            time.sleep(1)
            true_format = today_to_past.strftime("%Y/%m/%d")
            table, search_web, search_date = op.crawl(search_web='fu_three', search_date=true_format)
            print('Connecting to web...')
            sf_fucall = int(table[2].iloc[5][9]) #自營商大台
            sf_fuput = int(table[2].iloc[5][11])
            fucall = int(table[2].iloc[7][9])   #外資大台
            fuput = int(table[2].iloc[7][11])

            fu_money = int(table[2].iloc[7][14])#外資未平倉多空淨額
            sffu_money = int(table[2].iloc[5][14])#自營未平倉多空淨額

            sf_smfucall = int(table[2].iloc[14][9]) #自營商小台
            sf_smfuput = int(table[2].iloc[14][11])
            smfucall = int(table[2].iloc[16][9])  #外資小台
            smfuput = int(table[2].iloc[16][11])
                           
            print(true_format + ' 寫入檔案...')
            op_data.append([search_date, sf_fucall, sf_fuput, fucall, fuput, sf_smfucall, sf_smfuput, smfucall, smfuput, fu_money, sffu_money])
            if i == get_days:
                workbook = openpyxl.load_workbook('test.xlsx')
                sheet = workbook.worksheets[1]
                # pd_data = pd.Series(sheet)
                # print(pd_data)
                # sheet.append(['日期', '自營大台多', '自營大台空', '外資大台多', '外資大台空', '自營小台多', '自營小台空', '外資小台多', '外資小台空', '外資多空淨額', '自營多空淨額'])
                for data in op_data:
                    sheet.append(data)                    
                workbook.save('test.xlsx')
                break
        except:
            print(true_format)
            print('可能是假日或沒開盤')
            if i == get_days:
                workbook = openpyxl.load_workbook('test.xlsx')
                sheet = workbook.worksheets[1]
                # sheet.append(['日期', '自營大台多', '自營大台空', '外資大台多', '外資大台空', '自營小台多', '自營小台空', '外資小台多', '外資小台空', '外資多空淨額', '自營多空淨額'])
                for data in op_data:
                    sheet.append(data)
                workbook.save('test.xlsx')
                break

def big10fu():
    get_days = int(input('請輸入十大fu回測天數: '))
    i = 0
    op_data = []
    while True:
        try:
            day_delta = i
            i += 1
            today_to_past = (dt.datetime.now() - timedelta(days= int(day_delta)))
            time.sleep(1)
            true_format = today_to_past.strftime("%Y/%m/%d")

            table, search_web, search_date = op.crawl_big10(search_web='fu_big10', search_date=true_format)
            big10_call = table[3].iloc[2][5].split('%', 1)[0] #以%分割字串，分割1次,取出第一項
            big10_put = table[3].iloc[2][9].split('%', 1)[0]
            big5_call = table[3].iloc[2][3].split('%', 1)[0]
            big5_put = table[3].iloc[2][7].split('%', 1)[0]


                       
            print(true_format + '寫入檔案...')
            op_data.append([search_date, big10_call, big10_put, big5_call, big5_put])

            if i == get_days:
                workbook = openpyxl.load_workbook('test.xlsx')
                sheet = workbook.worksheets[4]
                # sheet.append(['日期', '十大期貨多單', '十大期貨空單', '五大期貨多單', '五大期貨空單']) 
                for data in op_data:
                    sheet.append(data)
                workbook.save('test.xlsx')
                break
        except:
            print(true_format)
            print('可能是假日或沒開盤或只有月選')
            if i == get_days:
                workbook = openpyxl.load_workbook('test.xlsx')
                sheet = workbook.worksheets[4]
                # sheet.append(['日期', '十大期貨多單', '十大期貨空單', '五大期貨多單', '五大期貨空單'])
                for data in op_data:
                    sheet.append(data)
                workbook.save('test.xlsx')
                break

def big10op_week():
    get_days = int(input('請輸入十大OP週選回測天數: '))
    i = 0
    op_data = []
    while True:
        try:
            day_delta = i
            i += 1
            today_to_past = (dt.datetime.now() - timedelta(days= int(day_delta)))
            time.sleep(1)
            true_format = today_to_past.strftime("%Y/%m/%d")

            table, search_web, search_date = op.crawl_big10(search_web='cnp_big10', search_date=true_format)
            big10_weekbc = int((table[3].iloc[0][4].split(' ', 1)[0]).replace(',','')) # 十大週選op #split, 用空白切割, 切割1次, 取出第一項
            big10_weekbp = int((table[3].iloc[0][8].split(' ', 1)[0]).replace(',',''))
            big10_weeksc = int((table[3].iloc[3][4].split(' ', 1)[0]).replace(',',''))
            big10_weeksp = int((table[3].iloc[3][8].split(' ', 1)[0]).replace(',',''))
            # print(table[3])

            # big10_monthbc = int((table[3].iloc[1][4].split(' ', 1)[0]).replace(',',''))# 十大月選op
            # big10_monthbp = int((table[3].iloc[1][8].split(' ', 1)[0]).replace(',',''))
            # big10_monthsc = int((table[3].iloc[4][4].split(' ', 1)[0]).replace(',',''))
            # big10_monthsp = int((table[3].iloc[4][8].split(' ', 1)[0]).replace(',',''))
                       
            print(true_format + '寫入檔案...')
            op_data.append([search_date, big10_weekbc, big10_weekbp, big10_weeksc, big10_weeksc])

            if i == get_days:
                workbook = openpyxl.load_workbook('test.xlsx')
                sheet = workbook.worksheets[2]
                # sheet.append(['日期', '週選bc', '週選bp', '週選sc', '週選sp']) 
                for data in op_data:
                    sheet.append(data)
                workbook.save('test.xlsx')
                break
        except:
            print(true_format)
            print('可能是假日或沒開盤或只有月選')
            if i == get_days:
                workbook = openpyxl.load_workbook('test.xlsx')
                sheet = workbook.worksheets[2]
                # sheet.append(['日期', '週選bc', '週選bp', '週選sc', '週選sp'])
                for data in op_data:
                    sheet.append(data)
                workbook.save('test.xlsx')
                break

def big10op_month():
    get_days = int(input('請輸入十大OP月選回測天數: '))
    i = 0
    op_data = []
    while True:
        try:
            day_delta = i
            i += 1
            today_to_past = (dt.datetime.now() - timedelta(days= int(day_delta)))
            time.sleep(1)
            true_format = today_to_past.strftime("%Y/%m/%d")

            table, search_web, search_date = op.crawl_big10(search_web='cnp_big10', search_date=true_format)
            # big10_weekbc = int((table[3].iloc[0][4].split(' ', 1)[0]).replace(',','')) # 十大週選op #split, 用空白切割, 切割1次, 取出第一項
            # big10_weekbp = int((table[3].iloc[0][8].split(' ', 1)[0]).replace(',',''))
            # big10_weeksc = int((table[3].iloc[3][4].split(' ', 1)[0]).replace(',',''))
            # big10_weeksp = int((table[3].iloc[3][8].split(' ', 1)[0]).replace(',',''))
            # print(table[3])

            big10_monthbc = int((table[3].iloc[1][4].split(' ', 1)[0]).replace(',',''))# 十大月選op
            big10_monthbp = int((table[3].iloc[1][8].split(' ', 1)[0]).replace(',',''))
            big10_monthsc = int((table[3].iloc[4][4].split(' ', 1)[0]).replace(',',''))
            big10_monthsp = int((table[3].iloc[4][8].split(' ', 1)[0]).replace(',',''))
                       
            print(true_format + '寫入檔案...')
            op_data.append([search_date, big10_monthbc, big10_monthbp, big10_monthsc, big10_monthsp])

            if i == get_days:
                workbook = openpyxl.load_workbook('test.xlsx')
                sheet = workbook.worksheets[3]
                # sheet.append(['日期', '月選bc', '月選bp', '月選sc', '月選sp']) 
                for data in op_data:
                    sheet.append(data)
                workbook.save('test.xlsx')
                break
        except:
            print(true_format)
            print('可能是假日或沒開盤或只有月選')
            if i == get_days:
                workbook = openpyxl.load_workbook('test.xlsx')
                sheet = workbook.worksheets[3]
                # sheet.append(['日期', '月選bc', '月選bp', '月選sc', '月選sp'])
                for data in op_data:
                    sheet.append(data)
                workbook.save('test.xlsx')
                break

def crawl_0050():
    get_days = int(input('請輸入0050回測天數: '))
    i = 0
    op_data = []
    while True:
        try:
            day_delta = i
            i += 1
            today_to_past = (dt.datetime.now() - timedelta(days= int(day_delta)))
            time.sleep(1)
            true_format = today_to_past.strftime("%Y%m%d")

            op_bc_bp = 'https://www.twse.com.tw/fund/MI_QFIIS?response=json&date=' + true_format +'&selectType=0099P&_=1592550135988'
            headers = {'User-Agent': 'Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/81.0.4044.113 Mobile Safari/537.36'}
            res = requests.get(op_bc_bp,headers=headers,timeout = 5)
            print("Crawling website...")
            df = pd.DataFrame(res.json()['data'])
            df1 = float(df[df.iloc[:,1] == '元大台灣50'][7])
            df2 = float(df[df.iloc[:,1] == '元大台灣50反1'][7])
            print(true_format + ' get data...')
            op_data.append([true_format, df1, df2])
            if i == get_days:
                workbook = openpyxl.load_workbook('test.xlsx')
                sheet = workbook.worksheets[5]
                # sheet.append(['日期', '0050', '50反1'])
                for data in op_data:
                    sheet.append(data)                    
                workbook.save('test.xlsx')
                break
        except:
            print(true_format)
            print('Error cause holiday or not marketing day')
            if i == get_days:
                workbook = openpyxl.load_workbook('test.xlsx')
                sheet = workbook.worksheets[5]
                # sheet.append(['日期', '0050', '50反1'])
                for data in op_data:
                    sheet.append(data)
                workbook.save('test.xlsx')
                break




if __name__ == '__main__':
   
    while True:
        q = input('\n執行op資料抓取輸入\'1\', \n執行fu資料抓取輸入\'2\', \n執行十大op週選資料抓取輸入\'3\', \n執行十大op月選資料抓取輸入\'4\', \n執行十大fu資料抓取輸入\'5\', \n執行0050資料抓取輸入\'6\'\n離開輸入\'q\': ')
        if q == '1':
            print('執行op資料抓...')
            analysis_op_excel()
        elif q == '2':
             print('執行fu資料抓取...')
             analysis_fu_excel()
        elif q == '3':
            print('執行十大op週選資料抓取...')
            big10op_week()
        elif q == '4':
            print('執行十大op月選資料抓取...')
            big10op_month()
        elif q == '5':
            print('執行十大fu資料抓取...')
            big10fu()            
        elif q == '6':
            print('執行0050資料抓取...')
            crawl_0050()     
        elif q == 'q':
            break